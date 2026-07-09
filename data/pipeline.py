# /// script
# requires-python = ">=3.12"
# dependencies = ["numpy>=2", "openpyxl>=3.1"]
# ///
"""Build src/data/states.json for the SNAP payment error rate tool.

Inputs (all official, committed under data/sources/ except the QC PUF):
  - fy{17,18,19,22,23,24,25}-per.txt — pdftotext -layout extracts of the FNS/FNA
    "SNAP: Payment Error Rates" tables (PDFs committed alongside). FY2020-21
    were not published (COVID QC waivers); FY2014-16 preceded the QC
    integrity overhaul and are not comparable.
  - snap-fy25-state-activity.xlsx — FNA National Data Bank monthly state
    participation/benefit summary, FY2025 (P-EBT excluded), for benefit
    issuance by state.
  - SNAP QC public-use file FY2024 (downloaded on demand from snapqcdata.net,
    sha256-pinned, cached in data/cache/) for sampling-error estimation.

Outputs:
  - src/data/states.json — one record per state agency with the official PER
    series, sampling SEs, OBBBA cost-share bins and probabilities, dollar
    exposure, sample-size counterfactuals, and signal-vs-noise decomposition.

Statutory basis (Food and Nutrition Act §4(a), as amended by OBBBA §10105,
Pub. L. 119-21; 7 U.S.C. 2013(a)(2)):
  (B)(i)  state share of benefit cost: <6% -> 0%; 6-<8% -> 5%; 8-<10% -> 10%;
          >=10% -> 15%, beginning FY2028.
  (B)(ii) FY2028 uses the state-elected FY2025 OR FY2026 PER; FY2029+ uses the
          PER of the third preceding fiscal year.
  (B)(iii) implementation is delayed to FY2029 (FY2030) for states whose
          FY2025 (FY2026) PER x 1.5 >= 20% — i.e. PER >= 13.33%.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import statistics
import sys
import urllib.request
import zipfile
from pathlib import Path

import numpy as np
import openpyxl

HERE = Path(__file__).parent
SOURCES = HERE / "sources"
CACHE = HERE / "cache"
OUT = HERE.parent / "src" / "data" / "states.json"

YEARS = [2017, 2018, 2019, 2022, 2023, 2024, 2025]
RECENT_YEARS = [2022, 2023, 2024, 2025]  # same-regime years for noise checks

# FY2024 QC tolerance threshold (errors below this are excluded from the PER).
# FY2025 is $57, FY2026 is $58; the PUF we bootstrap is FY2024.
TOLERANCE_FY2024 = 56.0

QC_PUF_URL = "https://snapqcdata.net/sites/default/files/2026-05/qcfy2024_csv.zip"
QC_PUF_SHA256 = "0f3230a4318307d3088382546095eebfde03e781da6f65c9eac7f077bd4263f4"
QC_PUF_MEMBER = "qc_pub_fy2024.csv"

DELAY_THRESHOLD = 20.0 / 1.5  # PER x 1.5 >= 20%  <=>  PER >= 13.333...%

BOOTSTRAP_REPS = 1000
SEED = 20260709

FIPS = {
    "01": "Alabama", "02": "Alaska", "04": "Arizona", "05": "Arkansas",
    "06": "California", "08": "Colorado", "09": "Connecticut", "10": "Delaware",
    "11": "District of Columbia", "12": "Florida", "13": "Georgia", "66": "Guam",
    "15": "Hawaii", "16": "Idaho", "17": "Illinois", "18": "Indiana",
    "19": "Iowa", "20": "Kansas", "21": "Kentucky", "22": "Louisiana",
    "23": "Maine", "24": "Maryland", "25": "Massachusetts", "26": "Michigan",
    "27": "Minnesota", "28": "Mississippi", "29": "Missouri", "30": "Montana",
    "31": "Nebraska", "32": "Nevada", "33": "New Hampshire", "34": "New Jersey",
    "35": "New Mexico", "36": "New York", "37": "North Carolina",
    "38": "North Dakota", "39": "Ohio", "40": "Oklahoma", "41": "Oregon",
    "42": "Pennsylvania", "44": "Rhode Island", "45": "South Carolina",
    "46": "South Dakota", "47": "Tennessee", "48": "Texas", "49": "Utah",
    "50": "Vermont", "78": "Virgin Islands", "51": "Virginia",
    "53": "Washington", "54": "West Virginia", "55": "Wisconsin", "56": "Wyoming",
}
NAME_TO_ABBREV = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Guam": "GU",
    "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN",
    "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA",
    "Maine": "ME", "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI",
    "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO", "Montana": "MT",
    "Nebraska": "NE", "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ",
    "New Mexico": "NM", "New York": "NY", "North Carolina": "NC",
    "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR",
    "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virgin Islands": "VI", "Virginia": "VA",
    "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}
CANONICAL = set(NAME_TO_ABBREV)


def canonical_name(raw: str) -> str | None:
    squashed = re.sub(r"\s+", " ", raw).strip().upper()
    if squashed in ("DIST. OF COL.", "DISTRICT OF COLUMBIA"):
        return "District of Columbia"
    name = squashed.title()
    return name if name in CANONICAL else None


ROW = re.compile(
    r"^\s*([A-Za-z][A-Za-z. ]+?)\s{2,}\*?(\d+\.\d+)\s+\*?(\d+\.\d+)\s+\*?(\d+\.\d+)\s*$"
)


def parse_per_table(path: Path) -> tuple[dict[str, dict[str, float]], dict[str, float]]:
    """Parse one pdftotext extract -> ({state: {over, under, per}}, national)."""
    states: dict[str, dict[str, float]] = {}
    national: dict[str, float] = {}
    for line in path.read_text().splitlines():
        m = ROW.match(line)
        if not m:
            continue
        raw, over, under, per = m.group(1), *map(float, m.group(2, 3, 4))
        if re.sub(r"\s+", " ", raw).strip().upper() in ("UNITED STATES", "TOTAL", "NATIONAL"):
            national = {"over": over, "under": under, "per": per}
            continue
        name = canonical_name(raw)
        if name:
            states[name] = {"over": over, "under": under, "per": per}
    if not national:
        raise ValueError(f"{path.name}: national row not found")
    if len(states) != 53:
        missing = CANONICAL - set(states)
        raise ValueError(f"{path.name}: parsed {len(states)} states; missing {sorted(missing)}")
    return states, national


def load_issuance() -> tuple[dict[str, float], float]:
    """FY2025 benefit issuance by state from the National Data Bank workbook."""
    wb = openpyxl.load_workbook(SOURCES / "snap-fy25-state-activity.xlsx", read_only=True)
    issuance: dict[str, float] = {}
    for sheet in wb.sheetnames:
        if sheet == "US Summary":
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        current: str | None = None
        for row in rows:
            first = row[0]
            if isinstance(first, str):
                label = first.strip()
                name = canonical_name(label)
                if name and row[1] is None:
                    current = name
                    continue
                if label == "Total" and current and isinstance(row[3], (int, float)):
                    issuance[current] = float(row[3])
                    current = None
    total = sum(v for k, v in issuance.items())
    if len(issuance) != 53:
        missing = CANONICAL - set(issuance)
        raise ValueError(f"issuance: got {len(issuance)}; missing {sorted(missing)}")
    return issuance, total


def ensure_qc_csv() -> Path:
    """Download (pinned), verify, extract, and cache the FY2024 QC PUF."""
    CACHE.mkdir(parents=True, exist_ok=True)
    csv_path = CACHE / QC_PUF_MEMBER
    if csv_path.exists():
        return csv_path
    zip_path = CACHE / "qcfy2024_csv.zip"
    if not zip_path.exists():
        print("downloading FY2024 QC PUF (~90 MB)...", file=sys.stderr)
        req = urllib.request.Request(
            QC_PUF_URL,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
                "Referer": "https://snapqcdata.net/datafiles",
            },
        )
        with urllib.request.urlopen(req) as resp:
            zip_path.write_bytes(resp.read())
    digest = hashlib.sha256(zip_path.read_bytes()).hexdigest()
    if digest != QC_PUF_SHA256:
        raise ValueError(f"QC PUF sha256 mismatch: {digest}")
    with zipfile.ZipFile(zip_path) as zf, zf.open(QC_PUF_MEMBER) as member:
        csv_path.write_bytes(member.read())
    return csv_path


def qc_sampling_error() -> dict[str, dict[str, float]]:
    """Per-state QC sample size and bootstrap SE of the dollar-weighted PER.

    Replicates the PER's form on the public FY2024 QC file: cases with a
    recorded error of at least the $56 FY2024 tolerance contribute the full
    error amount (over- and under-issuance both count positively, no netting);
    the denominator is reported issuance. The public file excludes ineligible
    and incomplete cases and carries no regression adjustment, so its level
    differs from the official PER; we use it for *relative* precision (CV),
    applied to the official rates.
    """
    by_state: dict[str, dict[str, list[float]]] = {}
    with open(ensure_qc_csv(), newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        i_state = header.index("STATE")
        i_amterr = header.index("AMTERR")
        i_rawben = header.index("RAWBEN")
        i_hwgt = header.index("HWGT")
        for row in reader:
            fips = row[i_state].strip().zfill(2)
            name = FIPS.get(fips)
            if name is None:
                raise ValueError(f"unknown STATE fips {fips!r}")
            err = float(row[i_amterr] or 0.0)
            ben = float(row[i_rawben] or 0.0)
            w = float(row[i_hwgt] or 0.0)
            rec = by_state.setdefault(name, {"err": [], "ben": [], "w": []})
            rec["err"].append(err if err >= TOLERANCE_FY2024 else 0.0)
            rec["ben"].append(ben)
            rec["w"].append(w)

    rng = np.random.default_rng(SEED)
    out: dict[str, dict[str, float]] = {}
    for name, rec in by_state.items():
        err = np.array(rec["err"])
        ben = np.array(rec["ben"])
        w = np.array(rec["w"])
        n = len(err)
        point = float((w * err).sum() / (w * ben).sum()) * 100
        idx = rng.integers(0, n, size=(BOOTSTRAP_REPS, n))
        num = (w[idx] * err[idx]).sum(axis=1)
        den = (w[idx] * ben[idx]).sum(axis=1)
        reps = num / den * 100
        se = float(reps.std(ddof=1))
        out[name] = {
            "n": n,
            "pufPer": round(point, 2),
            "bootSePpt": round(se, 3),
            "cv": se / point if point > 0 else float("nan"),
        }
    return out


SHARE_BINS = [
    ("lt6", 0.0, 6.0, 0.00),
    ("b6to8", 6.0, 8.0, 0.05),
    ("b8to10", 8.0, 10.0, 0.10),
    ("b10to13_33", 10.0, DELAY_THRESHOLD, 0.15),
    ("gte13_33", DELAY_THRESHOLD, math.inf, 0.15),
]


def share_for(per: float) -> float:
    if per < 6:
        return 0.0
    if per < 8:
        return 0.05
    if per < 10:
        return 0.10
    return 0.15


def normal_cdf(x: float) -> float:
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def bin_probabilities(per: float, se: float) -> dict[str, float]:
    """P(measured rate falls in each bin) if the true rate equals `per`.

    Diane Schanzenbach's framing: 'If the observed PER were correct, sampling
    error alone can predict likelihood of falling into different payment
    bins.' Normal approximation, truncated below at 0.
    """
    if se <= 0:
        return {key: (1.0 if lo <= per < hi else 0.0) for key, lo, hi, _ in SHARE_BINS}
    z0 = normal_cdf((0 - per) / se)  # mass below zero, renormalized away
    probs = {}
    for key, lo, hi, _ in SHARE_BINS:
        p_lo = normal_cdf((lo - per) / se)
        p_hi = 1.0 if hi is math.inf else normal_cdf((hi - per) / se)
        probs[key] = max(0.0, (p_hi - p_lo) / (1 - z0))
    total = sum(probs.values())
    return {k: v / total for k, v in probs.items()}


def cost_views(per: float, se: float, issuance: float, delayed: bool):
    """Point and probability-weighted FY2028 exposure, plus steady-state."""
    probs = bin_probabilities(per, se)
    point_share = share_for(per)
    point_fy2028 = 0.0 if delayed else point_share * issuance
    # FY2028 expected view mirrors Diane's table: the >=13.33 band carries a
    # temporary 0% share because implementation is delayed for those draws.
    expected_fy2028 = sum(
        (0.0 if key == "gte13_33" else share) * probs[key] * issuance
        for key, _, _, share in SHARE_BINS
    )
    steady_expected = sum(share * probs[key] * issuance for key, _, _, share in SHARE_BINS)
    return probs, point_fy2028, expected_fy2028, point_share * issuance, steady_expected


def main() -> None:
    tables = {y: parse_per_table(SOURCES / f"fy{str(y)[2:]}-per.txt") for y in YEARS}
    national_series = {y: tables[y][1] for y in YEARS}
    issuance, issuance_total = load_issuance()
    qc = qc_sampling_error()

    # National sanity checks against the published PDFs.
    assert national_series[2024]["per"] == 10.93, national_series[2024]
    assert national_series[2025]["per"] == 10.62, national_series[2025]

    states_out = []
    for name in sorted(CANONICAL):
        series = {y: tables[y][0].get(name, {}).get("per") for y in YEARS}
        per25 = series[2025]
        per24 = series[2024]
        rec25 = tables[2025][0][name]
        q = qc[name]
        se25 = q["cv"] * per25 if per25 else float("nan")
        delayed = per25 * 1.5 >= 20.0
        probs, point_fy2028, expected_fy2028, point_steady, expected_steady = cost_views(
            per25, se25, issuance[name], delayed
        )
        point_bin_key = next(
            key for key, lo, hi, _ in SHARE_BINS if lo <= per25 < hi
        )
        flip_prob = 1 - probs[point_bin_key]

        # Sampling-size counterfactuals: SE scales with 1/sqrt(k).
        sample_scenarios = []
        for k in (0.5, 1.0, 2.0, 4.0):
            se_k = se25 / math.sqrt(k)
            p_k, _, exp_k, _, _ = cost_views(per25, se_k, issuance[name], delayed)
            sample_scenarios.append(
                {
                    "k": k,
                    "sePpt": round(se_k, 3),
                    "binProbs": {b: round(v, 4) for b, v in p_k.items()},
                    "expectedFY2028": round(exp_k),
                    "flipProb": round(1 - p_k[point_bin_key], 4),
                }
            )

        # Signal vs noise: recent-regime year-to-year spread vs sampling SE.
        recent = [series[y] for y in RECENT_YEARS if series[y] is not None]
        sd_recent = statistics.stdev(recent) if len(recent) >= 3 else None
        mean_recent = statistics.fmean(recent) if recent else None
        se_at_mean = q["cv"] * mean_recent if mean_recent else None

        states_out.append(
            {
                "name": name,
                "abbrev": NAME_TO_ABBREV[name],
                "series": {str(y): series[y] for y in YEARS},
                "fy2025": {
                    "over": rec25["over"],
                    "under": rec25["under"],
                    "per": per25,
                },
                "fy2024Per": per24,
                "issuanceFY2025": round(issuance[name]),
                "qc": {"n": q["n"], "pufPer": q["pufPer"], "bootSePpt": q["bootSePpt"]},
                "sePpt": round(se25, 3),
                "ci95": [round(per25 - 1.96 * se25, 2), round(per25 + 1.96 * se25, 2)],
                "share": share_for(per25),
                "delayedFY2028": delayed,
                "binProbs": {b: round(v, 4) for b, v in probs.items()},
                "flipProb": round(flip_prob, 4),
                "pointCostFY2028": round(point_fy2028),
                "expectedCostFY2028": round(expected_fy2028),
                "pointCostSteadyState": round(point_steady),
                "expectedCostSteadyState": round(expected_steady),
                "sampleScenarios": sample_scenarios,
                "yoy": {
                    "recentYears": RECENT_YEARS,
                    "sdRecent": round(sd_recent, 3) if sd_recent is not None else None,
                    "seAtMeanRecent": round(se_at_mean, 3) if se_at_mean else None,
                    "noiseRatio": round(sd_recent / se_at_mean, 2)
                    if sd_recent is not None and se_at_mean
                    else None,
                },
            }
        )

    total_point = sum(s["pointCostFY2028"] for s in states_out)
    total_expected = sum(s["expectedCostFY2028"] for s in states_out)
    total_steady = sum(s["pointCostSteadyState"] for s in states_out)
    delayed_states = [s["abbrev"] for s in states_out if s["delayedFY2028"]]

    out = {
        "meta": {
            "generated": "2026-07-09",
            "statute": "7 U.S.C. 2013(a)(2), as added by Pub. L. 119-21 §10105 (OBBBA)",
            "bins": [
                {"range": "<6%", "share": 0.0},
                {"range": "6–<8%", "share": 0.05},
                {"range": "8–<10%", "share": 0.10},
                {"range": "≥10%", "share": 0.15},
            ],
            "delayRule": "PER × 1.5 ≥ 20% (≥13.33%) in FY2025 delays a state's first cost-share year to FY2029; in FY2026, to FY2030 (7 U.S.C. 2013(a)(2)(B)(iii)).",
            "electionRule": "FY2028 uses the state's FY2025 or FY2026 PER at the state's election; FY2029+ uses the third preceding fiscal year (7 U.S.C. 2013(a)(2)(B)(ii)).",
            "perSource": "FNS/FNA SNAP payment error rate tables, FY2017–19 and FY2022–25 (FY2020–21 not published under COVID QC waivers).",
            "issuanceSource": "FNA National Data Bank, SNAP monthly state benefit summary FY2025 (P-EBT excluded), data as of 2026-06-12.",
            "seSource": "Bootstrap (B=1000) of the dollar-weighted error ratio on the FY2024 SNAP QC public-use file (n=44,891), FY2024 $56 tolerance applied; state CVs scaled to official FY2025 rates.",
            "toleranceNote": "Errors under the QC tolerance threshold ($56 FY2024, $57 FY2025, $58 FY2026) are excluded from the PER.",
            "qcPufSha256": QC_PUF_SHA256,
        },
        "national": {
            "series": {str(y): national_series[y]["per"] for y in YEARS},
            "fy2025": national_series[2025],
            "issuanceFY2025": round(issuance_total),
            "totalPointCostFY2028": total_point,
            "totalExpectedCostFY2028": total_expected,
            "totalPointCostSteadyState": total_steady,
            "delayedStatesFY2028": delayed_states,
        },
        "states": states_out,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, indent=1) + "\n")

    # Console validation summary.
    print(f"states: {len(states_out)}; FY25 issuance total: ${issuance_total/1e9:.1f}B")
    print(f"FY2028 point cost (delayed states at $0): ${total_point/1e9:.2f}B")
    print(f"FY2028 expected cost (noise-weighted):    ${total_expected/1e9:.2f}B")
    print(f"Steady-state point cost (no delay):        ${total_steady/1e9:.2f}B")
    print(f"delayed via FY2025 rate: {delayed_states}")
    for ab in ("CA", "NY", "AZ", "TX", "AK", "SD", "VI"):
        s = next(x for x in states_out if x["abbrev"] == ab)
        print(
            f"  {ab}: per25={s['fy2025']['per']:5.2f} se={s['sePpt']:.2f} "
            f"n={s['qc']['n']:5d} pufPer={s['qc']['pufPer']:5.2f} "
            f"flip={s['flipProb']:.0%} cost=${s['pointCostFY2028']/1e6:.0f}M "
            f"exp=${s['expectedCostFY2028']/1e6:.0f}M yoyRatio={s['yoy']['noiseRatio']}"
        )


if __name__ == "__main__":
    main()
